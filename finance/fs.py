#! /usr/bin/python3
import re
import sys

import openpyxl
import pdfplumber


def extract_others(pdf: pdfplumber.pdf.PDF, begin: int, keys: list, subjects: dict):
    j = 0
    for i in range(begin, begin + 6):
        page = pdf.pages[i]
        tables = page.extract_tables()
        for table in tables:
            n = len(table)
            for k in range(n):
                # print(table[k])
                try:
                    if table[k][0] is not None and \
                            (table[k][1] != "" or table[k][2] != "" or k == n - 1 or
                             table[k + 1][0] is not None):
                        if table[k][1] is None and table[k][2] is None:
                            table[k][0] = table[k - 2][0] + table[k][0]
                            table[k][1] = table[k - 1][1]
                            table[k][2] = table[k - 1][2]
                        # print(table[k])
                        if table[k][0].endswith(keys[j]):
                            if table[k][1]:
                                # print(j, keys[j], table[k][1])
                                subjects[keys[j]] = table[k][1]
                            j += 1
                            if j == len(keys):
                                return
                except IndexError:
                    pass


def extract_dividend(pdf: pdfplumber.pdf.PDF, begin: int, subjects: dict):
    for i in range(begin, begin + 2):
        page = pdf.pages[i]
        tables = page.extract_tables()
        for table in tables:
            n = len(table)
            for k in range(n):
                try:
                    if table[k][0] is not None and \
                            (table[k][1] != "" or table[k][2] != "" or k == n - 1 or
                             table[k + 1][0] is not None):
                        if table[k][1] is None and table[k][2] is None:
                            table[k][0] = table[k - 2][0] + table[k][0]
                            table[k][1] = table[k - 1][1]
                            table[k][2] = table[k - 1][2]
                        if re.match(r'20\d\d年', table[k][0]) and \
                                re.match(r'[0-9,.]+', table[k][1]):
                            # print(0, "现金分红金额", table[k][1])
                            subjects["现金分红金额"] = table[k][1]
                            return
                except (IndexError, TypeError):
                    pass


# extract balance, income and cash_flow
def extract_statements(pdf: pdfplumber.pdf.PDF, begin: int, end: int, subjects: dict):
    keys = [
        # 1. 合并资产负债表
        # 1.1 资产
        # 1.1.1 流动资产
        "货币资金", "交易性金融资产", "以公允价值计量且其变动计入当期损益的金融资产",
        "应收票据", "应收账款", "应收款项融资", "预付款项",
        # 1.1.2 非流动资产
        "债权投资", "可供出售金融资产", "其他债权投资", "持有至到期投资", "长期应收款", "长期股权投资",
        "其他权益工具投资", "其他非流动金融资产", "投资性房地产", "固定资产", "在建工程", "工程物资",
        # 1.1.3 资产总计
        "资产总计",
        # 1.2 负债和所有者权益
        # 1.2.1 流动负债
        "短期借款", "应付票据", "应付账款", "预收款项", "一年内到期的非流动负债",
        # 1.2.2 非流动负债
        "长期借款", "应付债券", "长期应付款",
        # 1.2.3 负债合计
        "负债合计",
        # 2 合并利润表
        # 2.1 营业总收入
        "其中：营业收入",
        # 2.2 营业总成本
        "其中：营业成本", "税金及附加", "销售费用", "管理费用", "研发费用", "财务费用",
        # 2.3 营业利润
        # 2.4 利润总额
        "四、利润总额（亏损总额以“－”号填列）",
        # 2.5 净利润
        "五、净利润（净亏损以“－”号填列）", "归属于母公司所有者的净利润",
        # 3 合并现金流量表
        # 3.1 经营活动产生的现金流量
        "销售商品、提供劳务收到的现金",
        "经营活动产生的现金流量净额",
        # 3.2 投资活动产生的现金流量
        "处置固定资产、无形资产和其他长期资产收回的现金净额",
        "购建固定资产、无形资产和其他长期资产支付的现金",
        "投资活动产生的现金流量净额",
        # 3.3 筹资活动产生的现金流量
        "分配股利、利润或偿付利息支付的现金",
        "筹资活动产生的现金流量净额",
        # 3.4 汇率变动对现金及现金等价物的影响
        # 3.5 现金及现金等价物净增加额
        "五、现金及现金等价物净增加额",
        # 3.6 期末现金及现金等价物余额
        "六、期末现金及现金等价物余额"]

    j = -1
    for i in range(begin, end + 1):
        page = pdf.pages[i]
        tables = page.extract_tables({"keep_blank_chars": True})
        for table in tables:
            if table[0][0] == "项目":
                j += 1
            if j % 2:
                continue
            # print("new table", j)
            n = len(table)
            for k in range(n):
                if table[k][0] is not None and \
                        (table[k][1] != "" or table[k][2] != "" or k == n - 1 or
                         table[k + 1][0] is not None):
                    if table[k][1] is None and table[k][2] is None:
                        table[k][0] = table[k - 2][0] + table[k][0]
                        table[k][1] = table[k - 1][1]
                        table[k][2] = table[k - 1][2]
                    # print(table[k][0], table[k][1])
                    if table[k][1]:
                        key = re.sub(r"[.\d]", "", table[k][0])
                        if key in keys:
                            # print(key, table[k][1])
                            subjects[key] = table[k][1]


def get_locations(pdf: pdfplumber.pdf.PDF) -> dict:
    tabs = ["六、主要会计数据和财务指标", "普通股现金分红情况表", "1、合并资产负债表",
            "6、母公司现金流量表", "现金流量表补充资料"]
    locs = {}
    j = 0
    for i in range(len(pdf.pages)):
        page = pdf.pages[i]
        text = page.extract_text()
        if re.search(tabs[j], text):
            locs[tabs[j]] = i
            j += 1
            if j == len(tabs):
                break
    print(locs)
    return locs


def get_subjects(pdf: pdfplumber.pdf.PDF) -> dict:
    account = ["加权平均净资产收益率"]
    cash_flow2 = [
        # 现金流量表补充资料
        "固定资产折旧、油气资产折耗、生产性生物资产折旧",
        "无形资产摊销"]

    locs = get_locations(pdf)
    subjects = {}
    extract_others(pdf, locs["六、主要会计数据和财务指标"], account, subjects)
    extract_dividend(pdf, locs["普通股现金分红情况表"], subjects)
    extract_statements(pdf, locs["1、合并资产负债表"], locs["6、母公司现金流量表"], subjects)
    extract_others(pdf, locs["现金流量表补充资料"], cash_flow2, subjects)
    return subjects


def put_subjects(wb: openpyxl.workbook.workbook.Workbook, subjects: dict, column: int):
    sheet = wb["报表数据—录入"]
    for i in range(2, sheet.max_row):    # skip the first row
        subject = sheet.cell(row=i, column=1).value
        if subject:
            # subject = re.sub(r"[.\d]", "", subject.strip())
            subject = subject.strip()
            if subject in subjects:
                c = sheet.cell(row=i, column=column)
                v = subjects[subject]
                if v.endswith("%"):
                    c.style = "Percent"
                    c.number_format = "0.00%"
                    c.value = float(v.strip('%')) / 100
                else:
                    c.number_format = "#,##,0.00"
                    c.value = float(re.sub(",", "", v))


def test():
    # test
    with pdfplumber.open("000895双汇_2018.pdf") as pdf:
        locs = get_locations(pdf)
        subjects = {}
        extract_dividend(pdf, locs["普通股现金分红情况表"], subjects)
        print(subjects)


def main():
    n = len(sys.argv)
    if n < 2:
        print('Usage: ' + sys.argv[0] + 'fs_name')
        sys.exit(1)
    sys.stdout.flush()

    # test()
    wb = openpyxl.load_workbook("template.xlsx")
    for i in range(5):
        fs = sys.argv[1] + '_' + str(i + 2015) + '.pdf'
        print(fs)
        with pdfplumber.open(fs) as pdf:
            subjects = get_subjects(pdf)
            put_subjects(wb, subjects, i + 2)   # skip the first column
    wb.save(sys.argv[1] + '.xlsx')


if __name__ == "__main__":
    main()
